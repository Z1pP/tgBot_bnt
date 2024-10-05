import openpyxl
from openpyxl.styles import Font, Alignment
from datetime import datetime, timedelta
import calendar
import pathlib

from loader import db
from core.enums import Period


def get_period_dates(period: str) -> tuple:
    today = datetime.today()
    data_list: list = []
    start_date: str = ""
    end_date: str = ""

    if period == Period.ALL:
        data_list = db.get_report_list_for_excel_all()
        if data_list:
            start_date = data_list[0][0]
            end_date = data_list[-1][0]
        else:
            start_date = end_date = "Нет данных"
    elif period == Period.WEEKLY:
        start_date = (today - timedelta(days=today.weekday())).strftime("%d.%m.%Y")
        end_date = (today + timedelta(days=6 - today.weekday())).strftime("%d.%m.%Y")
    elif period == Period.MONTHLY:
        start_date = today.replace(day=1).strftime("%d.%m.%Y")
        last_day = calendar.monthrange(today.year, today.month)[1]
        end_date = today.replace(day=last_day).strftime("%d.%m.%Y")

    if period != Period.ALL:
        data_list = db.get_report_list_for_excel_period(
            start_date=start_date, end_date=end_date
        )

    return start_date, end_date, data_list


def create_excel(period: str) -> str:
    start_date, end_date, data_list = get_period_dates(period)

    # Создаем новую книгу Excel и выбираем активный лист
    wb = openpyxl.Workbook()
    ws = wb.active

    # Задаем шрифт и выравнивание для заголовков
    title_font = Font(bold=True)
    alignment = Alignment(horizontal="center")

    # Заголовки для столбцов
    headers = [
        "ДАТА",
        "ИМЯ",
        "ОБР ЗАПРОСЫ",
        "ВЫСТ СЧЕТА",
        "ОПЛ СЧЕТА",
        "МАРЖА",
        "ВЫРУЧКА",
        "КОНВ ЗАПР В СЧЁТ",
        "КОНВ СЧЕТ В ОПЛ",
        "ПРОЦЕНТ НАЦЕНКИ",
    ]

    # Устанавливаем заголовки и форматирование
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = title_font
        cell.alignment = alignment

    # Заполняем данные
    for row_num, row_data in enumerate(data_list, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(
                row=row_num, column=col_num, value=value if value is not None else "-"
            )
            cell.alignment = alignment

            # Форматирование числовых значений
            if isinstance(value, (int, float)):
                cell.number_format = "#,##0.00"

    folder = "folder_to_reports"
    pathlib.Path(folder).mkdir(parents=True, exist_ok=True)

    file = f"{folder}/Отчет {start_date}-{end_date}.xlsx"
    # Сохраняем Excel-файл
    wb.save(filename=file)
    return file
